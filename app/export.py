"""
Export invoice items to CSV or XLSX using a layout (columns with field path + title + blank).
"""
import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from .layouts import rows_from_items


def iso_datetime_to_date(value: Any) -> str:
    """Return YYYY-MM-DD from ISO timestamps."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if len(text) >= 10:
        return text[:10]
    return text


def export_to_csv(
    items: list[dict],
    columns: list[dict],
    filepath: str | Path,
    expand_array_path: str | None = None,
) -> None:
    """Write items to CSV. If expand_array_path is set, one row per array element with parent fields repeated (same headers)."""
    filepath = Path(filepath)
    headers = [c.get("title") or c.get("fieldPath") or "Column" for c in columns]
    rows = rows_from_items(items, columns, expand_array_path)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def export_to_xlsx(
    items: list[dict],
    columns: list[dict],
    filepath: str | Path,
    expand_array_path: str | None = None,
) -> None:
    """Write items to XLSX. If expand_array_path is set, one row per array element with parent fields repeated (same headers)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX export. Install with: pip install openpyxl")

    headers = [c.get("title") or c.get("fieldPath") or "Column" for c in columns]
    rows = rows_from_items(items, columns, expand_array_path)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")
    ws.title = "Invoices"

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)


def default_export_filename(extension: str) -> str:
    """e.g. invoices_20240209_143022.csv"""
    return f"invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extension}"
